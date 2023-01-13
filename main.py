from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook
import pandas as pd
#--------------------------------
root= Tk()
root.geometry('950x550')
root.title('Markt tools for Building')
now=datetime.datetime.now()
date=now.strftime("%Y-%m-%d")
#-------------------------------
wb=Workbook()
ws=wb.active
ws.title='customer'
ws["A1"]='Name'
ws["B1"]='phone'
ws["C1"]='Total'
ws["D1"]='Date buy'
wb.save('customers.xlsx')
 
def save():
    name=E_n.get()
    phone=E_p.get()
    t_cost=E_total.get()
    time=E_date.get()
    
    excel = openpyxl.load_workbook('customers.xlsx')
    file = excel.active
    file.cell(column=1,row=file.max_row+1,value=name)
    file.cell(column=2,row=file.max_row,value=phone)
    file.cell(column=3,row=file.max_row,value=t_cost)
    file.cell(column=4,row=file.max_row,value=time)
    excel.save('customers.xlsx')
#--------------------------------         
menu={ 
      0:['منشار',35] ,              1:['صاروخ',40],
      2:['عربة رمل',40],            3:['خلاطة',60],
      4:['شنيور',80],               5:['عربية',200],
      6:['خوذة',70],                7:['شاكوش',20],
      8:['متر قياس',20],            9:['مفتاح عادة',30],
      10:['كماشة',30],              11:['قصافة',30],
      12:['كماشة مسمار',30],        13:['مفتاح فرنسي',30],
      14:['س.معجون',25],             15:['مطرقة',40],
      16:['مفكات',20],               17:['م.اركت',80],
      18:['رافعة',120],              19:['حفارة',120],
      20:['حاجز بناء',50],           21:['مشبك',35],
      22:['مسامير',10],               23:['رافعة',100],
      24:['فارة',45]
      }
#-------------------------------------------
def bill():
    global E_n
    global E_date
    global E_p
    global E_total
    root.geometry('1205x552')
    F4=Frame(root,bg='#5f7161',width=250,height=434,bd=2,relief=GROOVE)
    F4.place(x=950,y=1)
    #------------------------------------------------------------------
    L_n=Label(F4,text='customer name:',bg='#5f7161',fg='white')
    L_n.place(x=20,y=5)
    E_n=Entry(F4,width=24,font=('normal',12),justify=CENTER)
    E_n.place(x=18,y=30)
    L_p=Label(F4,text='phone: ',bg='#5f7161',fg='white')
    L_p.place(x=20,y=55)
    E_p=Entry(F4,width=24,font=('normal',12),justify=CENTER)
    E_p.place(x=18,y=75)
    L_total=Label(F4,text='total price: ',bg='#5f7161',fg='white')
    L_total.place(x=20,y=100)
    E_total=Entry(F4,width=24,font=('normal',12),justify=CENTER)
    E_total.place(x=18,y=120)
    L_date=Label(F4,text='date of purchase: ',bg='#5f7161',fg='white')
    L_date.place(x=20,y=145)
    E_date=Entry(F4,width=24,font=('normal',12),justify=CENTER)
    E_date.place(x=18,y=165)
    
    add_b=Button(F4,text='Save the bill',width=30,cursor='hand2',bg='#EDDBC0',command=save)
    add_b.place(x=19 , y =195)
    
    add_b1=Button(F4,text='Empty fields',width=30,cursor='hand2',bg='#EDDBC0',command=clear1)
    add_b1.place(x=19 , y =225)
    #-----------------------------------------------------------------------------------
    total=0
    for item in trv.get_children():
        trv.delete(item)
    for i in range (len(listComponent)):
        if(int(listComponent[i].get())>0):
            price=int(listComponent[i].get())*menu[i][1]
            total=total+price
            myst=(str(menu[i][1]),str(listComponent[i].get()),str(price))
            trv.insert("",'end',iid=i,text=menu[i][0],values=myst)
    final=total
    E_total.insert('1',str(final)+'EGP')
    E_date.insert('1',str(date))           
#-------------------------------------------------------------------------------------------------
def clear():
    for item in trv.get_children():
        trv.delete(item)
    E_n.delete('0',END)
    E_p.delete('0',END)
    E_date.delete('0',END)
    E_total.delete('0',END)    
#-------------------------------------------------------------------------------------------------
def clear1():
    E_n.delete('0',END)
    E_p.delete('0',END)
     
def close():
    exit()
    
F1=Frame(root, bg='silver',width=600,height=550)
F1.place(x=1,y=1)
img_menu1=PhotoImage(file='img/1.png')
img_menu2=PhotoImage(file='img/2.png')
img_menu3=PhotoImage(file='img/3.png')
img_menu4=PhotoImage(file='img/4.png')
img_menu5=PhotoImage(file='img/5.png')
img_menu6=PhotoImage(file='img/6.png')
img_menu7=PhotoImage(file='img/7.png')
img_menu8=PhotoImage(file='img/8.png')
img_menu9=PhotoImage(file='img/9.png')
img_menu10=PhotoImage(file='img/10.png')
img_menu11=PhotoImage(file='img/11.png')
img_menu12=PhotoImage(file='img/12.png')
img_menu13=PhotoImage(file='img/13.png')
img_menu14=PhotoImage(file='img/14.png')
img_menu15=PhotoImage(file='img/15.png')
img_menu16=PhotoImage(file='img/16.png')
img_menu17=PhotoImage(file='img/17.png')
img_menu18=PhotoImage(file='img/18.png')
img_menu19=PhotoImage(file='img/19.png')
img_menu20=PhotoImage(file='img/20.png')
img_menu21=PhotoImage(file='img/21.png')
img_menu22=PhotoImage(file='img/22.png')
img_menu23=PhotoImage(file='img/23.png')
img_menu24=PhotoImage(file='img/24.png')
img_menu25=PhotoImage(file='img/25.png')

title=Label(F1,text='Construction equipment sales project',font=('Tajawal 10'),fg='white',bg='#61876E',width=75)
title.place(x=0,y=0)
menu1=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu1,text='منشار ',compound=TOP)
menu1.place(x=10,y=45)

menu2=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu2,text='صاروخ ',compound=TOP)
menu2.place(x=85,y=45)

menu3=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu3,text='عربةرمل',compound=TOP)
menu3.place(x=160,y=45)

menu4=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu4,text='خلاطة',compound=TOP)
menu4.place(x=237,y=45)

menu5=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu5,text='شنيور',compound=TOP)
menu5.place(x=310,y=45)

menu6=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu6,text='عربية',compound=TOP)
menu6.place(x=387,y=45)

menu7=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu7,text='خوذة',compound=TOP)
menu7.place(x=465,y=45)

menu8=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu8,text='شاكوش',compound=TOP)
menu8.place(x=10,y=145)

menu9=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu9,text='متر قياس',compound=TOP)
menu9.place(x=85,y=145)

menu10=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu10,text='مفتاح مقاس',compound=TOP)
menu10.place(x=160,y=145)

menu11=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu11,text='كماشة',compound=TOP)
menu11.place(x=237,y=145)

menu12=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu12,text='قصافة',compound=TOP)
menu12.place(x=310,y=145)

menu13=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu13,text='كماشةمسمار',compound=TOP)
menu13.place(x=387,y=145)

menu14=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu14,text='مفتاح فرنسي',compound=TOP)
menu14.place(x=465,y=145)

menu15=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu15,text='م.صمغ',compound=TOP)
menu15.place(x=10,y=240)

menu16=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu16,text='س.معجون',compound=TOP)
menu16.place(x=10,y=240)

menu17=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu17,text='مطرقة',compound=TOP)
menu17.place(x=85,y=240)

menu18=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu18,text='مفكات',compound=TOP)
menu18.place(x=160,y=240)

menu19=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu19,text='م.اركت',compound=TOP)
menu19.place(x=237,y=240)

menu20=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu20,text='مشبك',compound=TOP)
menu20.place(x=310,y=240)

menu21=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu21,text='حفارة',compound=TOP)
menu21.place(x=387,y=240)

menu22=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu22,text='حاجز بناء',compound=TOP)
menu22.place(x=465,y=240)

menu23=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu23,text='مسامير',compound=TOP)
menu23.place(x=10,y=340)

menu24=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu24,text='رافعة',compound=TOP)
menu24.place(x=85,y=340)

menu25=Button(F1,width=55,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=65,image=img_menu25,text='فارة',compound=TOP)
menu25.place(x=160,y=340)

listComponent=[]
font1=('Times',10,'bold')

com1=IntVar()
com2=IntVar()
com3=IntVar()
com4=IntVar()
com5=IntVar()
com6=IntVar()
com7=IntVar()
com8=IntVar()
com9=IntVar()
com10=IntVar()
com11=IntVar()
com12=IntVar()
com13=IntVar()
com14=IntVar()
com15=IntVar()
com16=IntVar()
com17=IntVar()
com18=IntVar()
com19=IntVar()
com20=IntVar()
com21=IntVar()
com22=IntVar()
com23=IntVar()
com24=IntVar()
com25=IntVar()


sb1=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com1)
sb1.place(x=10,y=120)
listComponent.append(sb1)

sb2=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com2)
sb2.place(x=85,y=120)
listComponent.append(sb2)

sb3=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com3)
sb3.place(x=160,y=120)
listComponent.append(sb3)

sb4=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com4)
sb4.place(x=237,y=120)
listComponent.append(sb4)

sb5=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com5)
sb5.place(x=310,y=120)
listComponent.append(sb5)

sb6=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com6)
sb6.place(x=387,y=120)
listComponent.append(sb6)

sb7=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com7)
sb7.place(x=465,y=120)
listComponent.append(sb7)

sb8=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com8)
sb8.place(x=10,y=218)
listComponent.append(sb8)

sb9=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com9)
sb9.place(x=85,y=218)
listComponent.append(sb9)

sb10=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com10)
sb10.place(x=160,y=218)
listComponent.append(sb10)

sb11=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com11)
sb11.place(x=237,y=218)
listComponent.append(sb11)

sb12=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com12)
sb12.place(x=310,y=218)
listComponent.append(sb12)

sb13=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com13)
sb13.place(x=387,y=218)
listComponent.append(sb13)

sb14=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com14)
sb14.place(x=465,y=218)
listComponent.append(sb14)

sb15=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com15)
sb15.place(x=10,y=318)
listComponent.append(sb15)

sb16=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com16)
sb16.place(x=85,y=318)
listComponent.append(sb16)

sb17=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com17)
sb17.place(x=160,y=318)
listComponent.append(sb17)

sb18=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com18)
sb18.place(x=237,y=318)
listComponent.append(sb18)

sb19=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com19)
sb19.place(x=310,y=318)
listComponent.append(sb19)

sb20=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com20)
sb20.place(x=387,y=318)
listComponent.append(sb20)

sb21=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com21)
sb21.place(x=465,y=318)
listComponent.append(sb21)

sb22=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com22)
sb22.place(x=310,y=318)
listComponent.append(sb22)

sb23=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com23)
sb23.place(x=10,y=418)
listComponent.append(sb23)

sb24=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com24)
sb24.place(x=85,y=418)
listComponent.append(sb24)

sb25=Spinbox(F1,from_=0,to_=5,font=font1,width=5,textvariable=com25)
sb25.place(x=160,y=418)
listComponent.append(sb25)

b1=Button(F1,text='Material purchase',fg='white',font=('Times 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command= bill)
b1.place(x=10,y=480)
b2=Button(F1,text='new bill',fg='white',font=('Times 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command=clear)
b2.place(x=150,y=480)
b4=Button(F1,text='Close the program',fg='white',font=('Times 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command=close)
b4.place(x=290,y=480)

F2=Frame(root,bg='#66CDAA',width=350,height=550)
F2.place(x=604 , y = 1)

trv=ttk.Treeview(F2 ,selectmode='browse')
trv.place(x=1 ,y = 1 ,width=340,height=550)

trv["columns"]=('1','2','3')
trv.column("#0",width=60,anchor='c')
trv.column("1",width=50,anchor='c')
trv.column("2",width=50,anchor='c')
trv.column("3",width=60,anchor='c')
trv.heading("#0",text="material",anchor='c')
trv.heading("1",text="price",anchor='c')
trv.heading("2",text="num",anchor='c')
trv.heading("3",text="total price",anchor='c')

root.mainloop()